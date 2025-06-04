import pandas as pd
import os
import io
import sys
import math
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
# from openpyxl import Workbook
from openpyxl.drawing.image import Image
import seaborn as sns


# def plot_function(directory_path: str):


def get_minimum_runtime_random(df):
    df['Filename'] = df['Filename'].apply(
        lambda x: '_'.join(x.split('/')[-1].split('_')[:2]))

    # Replace negative values with 0.00001
    df[' Run Time(ms)'] = df[' Run Time(ms)'].apply(lambda x: max(x, 0.01))

    pivot_df = df.pivot_table(values=' Run Time(ms)', index=[
        ' nn selection', ' iter selection'], columns='Filename')
    # Calculate the minimum runtime for each data file
    pivot_df.loc['Minimum Runtime'] = pivot_df.min(axis=0)
    pivot_df = pivot_df.rename_axis(columns=None)
    pivot_df.columns.name = 'Algorithm'
    pivot_df.name = "Run Time (ms) on Problem"

    return pivot_df


def get_minimum_runtime_greedy(df):
    df['Filename'] = df['Filename'].apply(
        lambda x: '_'.join(x.split('/')[-1].split('_')[:2]))

    # Replace negative values with 0.00001
    df[' Run Time(ms)'] = df[' Run Time(ms)'].apply(lambda x: max(x, 0.01))

    pivot_df = df.pivot_table(values=' Run Time(ms)', index=[
        ' Type of Algorithm'], columns='Filename')
    # Calculate the minimum runtime for each data file
    pivot_df.loc['Minimum Runtime'] = pivot_df.min(axis=0)
    pivot_df = pivot_df.rename_axis(columns=None)
    pivot_df.columns.name = 'Algorithm'
    pivot_df.name = "Run Time (ms) on Problem"

    return pivot_df


def get_max_profit_random(df):
    df['Filename'] = df['Filename'].apply(
        lambda x: '_'.join(x.split('/')[-1].split('_')[:2]))

    # Replace negative values with 0.00001
    df[' Profit after Local Search'] = df[' Profit after Local Search'].apply(
        lambda x: max(x, 0.01))

    pivot_df = df.pivot_table(values=' Profit after Local Search', index=[
        ' nn selection', ' iter selection'], columns='Filename')
    # Calculate the minimum runtime for each data file
    pivot_df.loc['Maximun Profit'] = pivot_df.max(axis=0)
    pivot_df = pivot_df.rename_axis(columns=None)
    pivot_df.columns.name = 'Algorithm'

    return pivot_df


def get_max_profit_greedy(df):
    df['Filename'] = df['Filename'].apply(
        lambda x: '_'.join(x.split('/')[-1].split('_')[:2]))

    # Replace negative values with 0.00001
    df[' Profit after Local Search'] = df[' Profit after Local Search'].apply(
        lambda x: max(x, 0.01))

    pivot_df = df.pivot_table(values=' Profit after Local Search', index=[
        ' Type of Algorithm'], columns='Filename')
    # Calculate the minimum runtime for each data file
    pivot_df.loc['Maximun Profit'] = pivot_df.max(axis=0)
    pivot_df = pivot_df.rename_axis(columns=None)
    pivot_df.columns.name = 'Algorithm'

    return pivot_df


def ratio_table(df):
    # Copy the pivot table to preserve the original
    ratio_df = df.copy()
    # Extract the minimum runtime values for each data file
    if 'Minimum Runtime' in ratio_df.index:
        min_value = ratio_df.loc['Minimum Runtime']
        ratio_df = ratio_df.drop('Minimum Runtime')
        for column in ratio_df.columns:
            if column in min_value.index:
                ratio_df[column] = np.log2(
                    ratio_df[column] / min_value[column])
    elif 'Maximun Profit' in ratio_df.index:
        max_value = ratio_df.loc['Maximun Profit']
        ratio_df = ratio_df.drop('Maximun Profit')
        for column in ratio_df.columns:
            if column in max_value.index:
                ratio_df[column] = np.log2(
                    max_value[column] / ratio_df[column])

    return ratio_df


def get_gap_table_random(df):
    df['Filename'] = df['Filename'].apply(
        lambda x: '_'.join(x.split('/')[-1].split('_')[:2]))

    pivot_df = df.pivot_table(values=' Error', index=[
        ' nn selection', ' iter selection'], columns='Filename')
    pivot_df = pivot_df.rename_axis(columns=None)
    pivot_df.columns.name = 'Algorithm'

    return pivot_df


def get_gap_table_greedy(df):
    df['Filename'] = df['Filename'].apply(
        lambda x: '_'.join(x.split('/')[-1].split('_')[:2]))

    pivot_df = df.pivot_table(values=' Error', index=[
        ' Type of Algorithm'], columns='Filename')
    pivot_df = pivot_df.rename_axis(columns=None)
    pivot_df.columns.name = 'Algorithm'

    return pivot_df


def number_instances_ratio_table(df):
    unique_values = np.unique(df.values)
    unique_values = np.sort(unique_values)
    number_instances_ratio_df = pd.DataFrame(
        index=df.index, columns=unique_values)
    # Iterate over rows and columns of original table
    for row_idx, row in df.iterrows():
        for col in number_instances_ratio_df.columns:
            # Count values <= current column value
            count = (row <= col).sum()
            # Assign count to new table
            number_instances_ratio_df.at[row_idx, col] = count

    return number_instances_ratio_df


def percentage_instances_ratio_table(minimum_df, number_instances_ratio_df):
    num_columns_minimum_df = len(minimum_df.columns)
    # print("num_columns_minimum_df: ", num_columns_minimum_df)
    percentage_instances_ratio_df = number_instances_ratio_df.copy()

    for column in percentage_instances_ratio_df.columns:
        percentage_instances_ratio_df[column] /= num_columns_minimum_df

    return percentage_instances_ratio_df


def stair_step_graph(df, type, graph_name):
    # Close any existing figures
    plt.close('all')
    plt.figure(figsize=(12, 6))
    x_values = [float(col) for col in df.columns]

    linestyles = ['-', '--', '-.', ':']

    min_yvalue = np.inf

    # Plot each row as a separate step graph
    for idx, row in enumerate(df.iterrows()):
        # color = dark_colors[idx % len(dark_colors)]
        linestyle_idx = idx % len(linestyles)
        plt.step(x_values, row[1], linewidth=1.5, where='post',
                 label=row[0], linestyle=linestyles[linestyle_idx])
        # Update minimum y value if needed
        min_yvalue = min(min_yvalue, min(row[1]))
        yvalue = "P((Ratio <= x))"
    else:
        yvalue = "P(log_10(Ratio <= x))"

    cutoff = 3 if max(x_values) > 3 else max(x_values)
    plt.yticks([i / 10 for i in range(11)])
    # plt.yticks([i / 10 for i in range(11)], [i / 10 for i in range(11)])
    plt.title(graph_name)
    plt.xlim(0, cutoff)
    plt.ylim(min_yvalue)
    plt.xlabel('Ratio of Algorithm vs. Data files')
    plt.ylabel(yvalue)
    plt.legend()
    plt.grid(True)
    # Save the plot to a buffer
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    plt.close()
    # Rewind the buffer
    buffer.seek(0)

    return buffer


def ratio_on_problem(minimun_maximun_df, worksheet, excel_writer):
    ratio_df = ratio_table(minimun_maximun_df)

    # Ratio on Problem
    ratio_df_endcol = openpyxl.utils.cell.get_column_letter(
        len(ratio_df.columns) + 1)
    ratio_df_start_row = len(minimun_maximun_df.index) + 4
    worksheet.merge_range(
        f'A{ratio_df_start_row}:{ratio_df_endcol}{ratio_df_start_row}', 'log_10(Ratio) on Problem')
    ratio_df.to_excel(excel_writer, sheet_name=worksheet.get_name(),
                      startrow=ratio_df_start_row, startcol=0)
    numberOfInstances_percentageInstancesRatio(
        ratio_df, ratio_df_start_row, worksheet, excel_writer, "ratio")


def numberOfInstances_percentageInstancesRatio(df, df_start_row, worksheet, excel_writer, type):
    sheetname = worksheet.get_name()
    number_instances_ratio_df = number_instances_ratio_table(df)
    percentage_instances_ratio_df = percentage_instances_ratio_table(
        df, number_instances_ratio_df)
    buffer = stair_step_graph(
        percentage_instances_ratio_df, type, graph_name=f'{sheetname} Graph')

    # Number of Instances with Ratio
    number_instances_ratio_df_endcol = openpyxl.utils.cell.get_column_letter(
        len(number_instances_ratio_df.columns) + 1)
    number_instances_ratio_df_start_row = len(
        df.index) + df_start_row + 3
    worksheet.merge_range(
        f'A{number_instances_ratio_df_start_row}:{number_instances_ratio_df_endcol}{number_instances_ratio_df_start_row}', 'Number of Instances with Ratio')
    number_instances_ratio_df.to_excel(
        excel_writer, sheet_name=sheetname, startrow=number_instances_ratio_df_start_row, startcol=0)

    # Percentage of Intances with Ratio
    percentage_instances_ratio_df_endcol = openpyxl.utils.cell.get_column_letter(
        len(percentage_instances_ratio_df.columns) + 1)
    percentage_instances_ratio_df_start_row = len(
        number_instances_ratio_df.index) + number_instances_ratio_df_start_row + 3
    worksheet.merge_range(
        f'A{percentage_instances_ratio_df_start_row}:{percentage_instances_ratio_df_endcol}{percentage_instances_ratio_df_start_row}', 'Percentage of Intances with Ratio')
    percentage_instances_ratio_df.to_excel(
        excel_writer, sheet_name=sheetname, startrow=percentage_instances_ratio_df_start_row, startcol=0)

    # Insert the image into the Excel file
    worksheet = excel_writer.sheets[sheetname]
    worksheet.insert_image(len(
        percentage_instances_ratio_df.index) + percentage_instances_ratio_df_start_row + 3, 0, '', {'image_data': buffer})


def minimun_runtime_data(type, df, excel_writer):
    # add Runtime sheet to the excel
    workbook = excel_writer.book
    worksheet = workbook.add_worksheet('Runtime')

    if (type == 'random'):
        minimum_df = get_minimum_runtime_random(df)
    elif (type == 'greedy'):
        minimum_df = get_minimum_runtime_greedy(df)

    # Runtime on Problem
    minimun_df_endcol = openpyxl.utils.cell.get_column_letter(
        len(minimum_df.columns) + 1)
    worksheet.merge_range(
        f'A1:{minimun_df_endcol}1', 'Run time (ms) on Problem')
    minimum_df.to_excel(excel_writer, sheet_name='Runtime',
                        startrow=1, startcol=0)

    ratio_on_problem(
        minimum_df, worksheet, excel_writer)


def maximum_profit_data(type, df, excel_writer):
    workbook = excel_writer.book
    worksheet = workbook.add_worksheet('Profit')

    if (type == 'random'):
        maximum_profit_df = get_max_profit_random(df)
    elif (type == 'greedy'):
        maximum_profit_df = get_max_profit_greedy(df)

    # Runtime on Problem
    maximum_profit_endcol = openpyxl.utils.cell.get_column_letter(
        len(maximum_profit_df.columns) + 1)
    worksheet.merge_range(
        f'A1:{maximum_profit_endcol}1', 'Profit on Problem')
    maximum_profit_df.to_excel(excel_writer, sheet_name=worksheet.get_name(),
                               startrow=1, startcol=0)

    ratio_on_problem(
        maximum_profit_df, worksheet, excel_writer)


def gap_on_problem(type, df, excel_writer):
    workbook = excel_writer.book
    worksheet = workbook.add_worksheet('Gap')

    if (type == 'random'):
        gap_df = get_gap_table_random(df)
    if (type == 'greedy'):
        gap_df = get_gap_table_greedy(df)

    # # Gap on Problem
    gap_df_endcol = openpyxl.utils.cell.get_column_letter(
        len(gap_df.columns) + 1)
    worksheet.merge_range(
        f'A1:{gap_df_endcol}1', 'Gap on Problem')
    gap_df.to_excel(excel_writer, sheet_name=worksheet.get_name(),
                    startrow=1, startcol=0)

    numberOfInstances_percentageInstancesRatio(
        gap_df, len(gap_df.index), worksheet, excel_writer, type="gap")


# directory_path = sys.argv[1]
# destination_path = sys.argv[2]

# destination_path = "../graphs/roulette-randomized/sparse_data_60%/disturb-numOfNode-0.7-7-3000ms-3nodeLeast"


def performance_profile(type, directory_path):
    all_csv_files = []

    csv_files = [file for file in os.listdir(
        directory_path) if file.endswith('.csv')]

    for file in csv_files:
        file_path = os.path.join(directory_path, file)
        df = pd.read_csv(file_path)
        all_csv_files.append(df)

    all_csv_df = pd.concat(all_csv_files, ignore_index=True)

    excel_writer = pd.ExcelWriter(os.path.join(
        directory_path, "Performance-Profile.xlsx"), engine='xlsxwriter')
    all_csv_df.to_excel(excel_writer, sheet_name="All Data", index=False)
    minimun_runtime_data(type, all_csv_df, excel_writer)
    maximum_profit_data(type, all_csv_df, excel_writer)
    gap_on_problem(type, all_csv_df, excel_writer)

    with excel_writer:
        pass


# Check if the directory_path argument is provided
if len(sys.argv) != 3:
    print("Usage: python3 performance-profile.py <directory_path> <algorithm-type>")
    sys.exit(1)

directory_path = sys.argv[-2]
# print(directory_path, sys.argv[-1])
performance_profile(sys.argv[-1], directory_path)
